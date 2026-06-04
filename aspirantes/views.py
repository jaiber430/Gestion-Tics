from django.shortcuts import render, redirect, get_object_or_404
from django.http import Http404

import os
from django.conf import settings
from aspirantes.utils import eliminar_carpetas_vencidas, combinar_pdfs

from django.core.files.base import ContentFile

# from .pdf_processor import process_pdf_image

# Importar datos de los modulos requeridos
from Cursos.models import Aspirantes, Solicitud, Caracterizacion, Tipoidentificacion

from datetime import datetime

# IMPORTS LOCALES PARA NO MODIFICAR EL TOPE DEL ARCHIVO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from django.contrib import messages
# Importar path desde pathlib para elimnar archivos (Forma moderna)
from pathlib import Path

def formulario_aspirantes(request, idsolicitud):

    solicitud = get_object_or_404(Solicitud, idsolicitud=idsolicitud)

    # Asegurarse de que no se puedan voler a registrar aspirantes una vez se cumpla la cantidad
    cerrar_inscripciones = Aspirantes.objects.filter(
        solicitudinscripcion=solicitud
    ).count()

    # Si se alcanza o supera el cupo → enviar a la página informativa
    if cerrar_inscripciones >= solicitud.cupo:
        # Mantener el link marcado como cerrado para evitar nuevas inscripciones,
        # pero sin mandar al 404: se muestra una página clara de "cupo agotado".
        if solicitud.linkpreinscripcion != 1:
            solicitud.linkpreinscripcion = 1
            solicitud.save(update_fields=['linkpreinscripcion'])
        return redirect('cupo_agotado')

    # Si el link fue desactivado manualmente (y aún hay cupo) → 404
    if solicitud.linkpreinscripcion == 1:
        raise Http404

    caracterizacion = Caracterizacion.objects.all()
    tipo_documento = Tipoidentificacion.objects.all()

    return render(request, 'forms/formulario_aspirantes.html', {
        'tipos_identificacion': tipo_documento,
        'caracterizaciones': caracterizacion,
        'solicitud': solicitud,
    })


def cupo_agotado(request):
    return render(request, 'pages/cupo_agotado.html', {
        'title': 'Cupos agotados',
    })

def registro_aspirante(request):
    if request.method == "POST":

        try:
            nombres = request.POST.get('nombres').upper()
            apellidos = request.POST.get('apellidos').upper()
            caracterizacion_id = request.POST.get('tipo_caracterizacion')
            telefono = int(request.POST.get('telefono'))
            pdf = request.FILES.get('pdf_documento')
            tipo_documento_id = request.POST.get('tipo_documento')
            identificacion = int(request.POST.get('numero_identificacion'))
            correo = request.POST.get('correo')
            solicitud_inscripcion = request.POST.get('idsolicitud')

            fecha_registro = datetime.now()

            # Bloquear registro si el cupo ya está completo (validación del lado servidor)
            try:
                solicitud_obj = Solicitud.objects.get(idsolicitud=solicitud_inscripcion)
            except Solicitud.DoesNotExist:
                messages.error(request, 'La solicitud no existe.')
                return redirect('formularioaspirantes', idsolicitud=solicitud_inscripcion)

            conteo_actual = Aspirantes.objects.filter(solicitudinscripcion=solicitud_obj).count()
            if conteo_actual >= solicitud_obj.cupo:
                if solicitud_obj.linkpreinscripcion != 1:
                    solicitud_obj.linkpreinscripcion = 1
                    solicitud_obj.save(update_fields=['linkpreinscripcion'])
                return redirect('cupo_agotado')

            # ✅ CORRECCIÓN: Validar duplicados SOLO dentro de la misma solicitud
            duplicado = Aspirantes.objects.filter(
                solicitudinscripcion=solicitud_obj,
                telefono=telefono
            ).exists() or Aspirantes.objects.filter(
                solicitudinscripcion=solicitud_obj,
                numeroidentificacion=identificacion
            ).exists() or Aspirantes.objects.filter(
                solicitudinscripcion=solicitud_obj,
                correo=correo
            ).exists()

            if duplicado:
                messages.error(request, 'Las credenciales ya han sido registradas en esta solicitud')
                return redirect('formularioaspirantes', idsolicitud=solicitud_inscripcion)

            # Obtener objetos relacionados
            id_tipo_caracterizacion = Caracterizacion.objects.get(idcaracterizacion=caracterizacion_id)
            id_tipo_documento = Tipoidentificacion.objects.get(idtipoidentificacion=tipo_documento_id)
            id_solicitud_preinscripcion = Solicitud.objects.get(idsolicitud=solicitud_inscripcion)

            # Crear carpeta con la solicitud para almacenar los pdf
            instructor = id_solicitud_preinscripcion

            # Subcarpeta con el id de la solicitud
            carpeta_solicitud = f"solicitud_{instructor.idsolicitud}"
            pdf_aspirantes = os.path.join(settings.MEDIA_ROOT, 'pdf', carpeta_solicitud)
            os.makedirs(pdf_aspirantes, exist_ok=True)

            # Nombre del archivo usando el número de identificación
            nombre_archivo = f"{identificacion}.pdf"
            direccion_archivo = os.path.join(pdf_aspirantes, nombre_archivo)

            # Validar que no exista ya un PDF con ese documento
            if os.path.exists(direccion_archivo):
                messages.error(request, 'Ya existe un aspirante registrado con este documento.')
                return redirect('formularioaspirantes', idsolicitud=solicitud_inscripcion)

            # Guardar PDF manualmente en la carpeta creada
            with open(direccion_archivo, 'wb+') as destino:
                for chunk in pdf.chunks():
                    destino.write(chunk)

            # Crear aspirante y asignar ruta relativa correcta en el campo pdf
            ruta_relativa_pdf = os.path.join('pdf', carpeta_solicitud, nombre_archivo)
            Aspirantes.objects.create(
                nombre=nombres,
                apellido=apellidos,
                idcaracterizacion=id_tipo_caracterizacion,
                telefono=int(telefono),
                pdf=ruta_relativa_pdf,
                tipoidentificacion=id_tipo_documento,
                numeroidentificacion=int(identificacion),
                correo=correo,
                fecha=fecha_registro,
                solicitudinscripcion=id_solicitud_preinscripcion,
            )

            # Primero ordenas los aspirantes
            aspirantes = Aspirantes.objects.filter(
                solicitudinscripcion=id_solicitud_preinscripcion
            ).order_by("idaspirante")

            # Luego haces el conteo
            total_aspirantes = aspirantes.count()

            if total_aspirantes >= id_solicitud_preinscripcion.cupo:
                combinar_pdfs(pdf_aspirantes)

                # ==============================================================
                # Generar archivo Excel con base en los aspirantes
                # ==============================================================

                solicitud = id_solicitud_preinscripcion

                programa = solicitud.codigoprograma
                nombre_programa = programa.nombreprograma

                nuevo_archivo = Workbook()
                hoja = nuevo_archivo.active
                hoja.title = f"Aspirantes Inscritos"

                hoja.merge_cells("A1:G1")
                celda = hoja["A1"]
                celda.value = "FORMATO PARA LA INSCRIPCIÓN DE ASPIRANTES EN SOFIA PLUS v1.0"

                celda.font = Font(bold=True, color="FFFFFF")
                celda.fill = PatternFill(start_color="66BB6A", end_color="66BB6A", fill_type="solid")
                celda.alignment = Alignment(horizontal="center", vertical="center")

                hoja.row_dimensions[1].height = 28.5

                encabezados = [
                    'Resultado del Registro(Reservado para el sistema)',
                    'Tipo de identificacion',
                    'Numero de identificacion',
                    'Código de ficha',
                    'Tipo población aspirantes',
                    '',
                    'Codigo empresa (solo si la ficha es cerrada)'
                ]
                hoja.append(encabezados)

                hoja.row_dimensions[2].height = 51

                for col_idx in range(1, len(encabezados) + 1):
                    cell = hoja.cell(row=2, column=col_idx)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                aspirantes = Aspirantes.objects.filter(
                    solicitudinscripcion=solicitud
                ).order_by("numeroidentificacion")

                for agregar in aspirantes:
                    caracterizacion = agregar.idcaracterizacion
                    tipo_caracterizacion = caracterizacion.caracterizacion if caracterizacion else ''

                    identificacion_tipo = agregar.tipoidentificacion
                    tipo_identificacion = identificacion_tipo.tipoidentificacion if identificacion_tipo else ''

                    empresa = agregar.solicitudinscripcion.idempresa
                    if empresa is not None:
                        mostrar_empresa = empresa.nitempresa
                    else:
                        mostrar_empresa = ''

                    hoja.append([
                        '',
                        tipo_identificacion,
                        agregar.numeroidentificacion,
                        '',
                        tipo_caracterizacion,
                        '',
                        mostrar_empresa
                    ])

                nombre_archivo_excel = f"formato_inscripcion_{solicitud.idsolicitud}.xlsx"
                carpeta_excel = os.path.join(settings.MEDIA_ROOT, 'excel')
                os.makedirs(carpeta_excel, exist_ok=True)

                directorio_excel = os.path.join(carpeta_excel, nombre_archivo_excel)

                nuevo_archivo.save(directorio_excel)

            messages.success(request, 'Te has preinscrito exitosamente')
            return redirect('formularioaspirantes', idsolicitud=solicitud_inscripcion)

        except Exception as e:
            messages.error(request, f'Error al registrarte: {e}')
            return redirect('formularioaspirantes', idsolicitud=request.POST.get('idsolicitud'))

    # GET request
    caracterizacion = Caracterizacion.objects.all()
    tipo_documento = Tipoidentificacion.objects.all()

    return render(request, 'forms/formulario_aspirantes.html', {
        'tipos_identificacion': tipo_documento,
        'caracterizaciones': caracterizacion,
    })

def updateCandidate(request, idSolicitud, numDoc):
    if request.method != "POST":
        raise Http404()

    aspirante = get_object_or_404(
        Aspirantes,
        numeroidentificacion=numDoc,
        solicitudinscripcion=idSolicitud
    )

    try:
        doc_antiguo = aspirante.numeroidentificacion
        doc_nuevo = int(request.POST.get('numero_identificacion'))
        nombre = request.POST.get('nombres').upper()
        apellido = request.POST.get('apellidos').upper()
        tipo_documento = int(request.POST.get('tipo_documento'))

        # ===============================
        # 1. ACTUALIZAR SOLO DB (CLAVE)
        # ===============================
        Aspirantes.objects.filter(
            idaspirante=aspirante.idaspirante
        ).update(
            nombre=nombre,
            apellido=apellido,
            tipoidentificacion_id=tipo_documento,
            numeroidentificacion=doc_nuevo
        )

        # ===============================
        # 2. RENOMBRAR PDF DEL ASPIRANTE
        # ===============================
        basePath = Path(settings.MEDIA_ROOT)
        folderPdf = basePath / "pdf" / f"solicitud_{idSolicitud}"
        pdfOld = folderPdf / f"{doc_antiguo}.pdf"
        pdfNew = folderPdf / f"{doc_nuevo}.pdf"

        if doc_antiguo != doc_nuevo and pdfOld.exists():
            pdfOld.rename(pdfNew)

        combinar_pdfs(folderPdf)

        # ===============================
        # 3. REGENERAR EXCEL
        # ===============================
        aspirantes = Aspirantes.objects.filter(
            solicitudinscripcion=idSolicitud
        ).order_by("numeroidentificacion")

        excelPath = basePath / "excel" / f"formato_inscripcion_{idSolicitud}.xlsx"
        excelPath.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "Aspirantes Inscritos"

        # ===== TÍTULO =====
        ws.merge_cells("A1:G1")
        celda = ws["A1"]
        celda.value = "FORMATO PARA LA INSCRIPCIÓN DE ASPIRANTES EN SOFIA PLUS v1.0"
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill(start_color="66BB6A", end_color="66BB6A", fill_type="solid")
        celda.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28.5

        # ===== ENCABEZADOS =====
        encabezados = [
            "Resultado del Registro (Reservado para el sistema)",
            "Tipo de identificación",
            "Número de identificación",
            "Código de ficha",
            "Tipo población aspirantes",
            "",
            "Código empresa (solo si la ficha es cerrada)"
        ]
        ws.append(encabezados)
        ws.row_dimensions[2].height = 51

        for col_idx in range(1, len(encabezados) + 1):
            cell = ws.cell(row=2, column=col_idx)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # ===== DATOS =====
        for asp in aspirantes:
            ws.append([
                "",
                asp.tipoidentificacion.tipoidentificacion if asp.tipoidentificacion else "",
                asp.numeroidentificacion,
                "",
                asp.idcaracterizacion.caracterizacion if asp.idcaracterizacion else "",
                "",
                asp.solicitudinscripcion.idempresa.nitempresa if asp.solicitudinscripcion.idempresa else ""
            ])

        wb.save(excelPath)
        messages.success(request, "Aspirante actualizado y Excel regenerado correctamente")

    except Exception as e:
        messages.error(request, f"Error al actualizar aspirante: {e}")

    return redirect('consultas_instructor')


# Eliminar aspirante
def removeApplicant(request, idSolicitud, numDoc):

    folderName = Path(settings.MEDIA_ROOT)

    try:
        # Eliminar pdf del aspirante
        deleteDocumentFile = folderName / "pdf" / f"solicitud_{idSolicitud}" / f"{numDoc}.pdf"
        if deleteDocumentFile.exists():
            deleteDocumentFile.unlink()

        # Eliminar aspirante de la DB
        eliminado, _ = Aspirantes.objects.filter(
            numeroidentificacion=numDoc,
            solicitudinscripcion=idSolicitud
        ).delete()

        if eliminado == 0:
            messages.warning(request, "No se encontró el aspirante para eliminar")
            return redirect('consultas_instructor')

        # Reactivar link de preinscripción (si se eliminó al menos un aspirante)
        Solicitud.objects.filter(idsolicitud=idSolicitud).update(
            linkpreinscripcion=0
        )

        # Eliminar pdf combinado
        deleteCombinedPdf = folderName / "pdf" / f"solicitud_{idSolicitud}" / "combinado.pdf"
        if deleteCombinedPdf.exists():
            deleteCombinedPdf.unlink()

        # Eliminar masivo excel
        deleteExcel = folderName / "excel" / f"formato_inscripcion_{idSolicitud}.xlsx"
        if deleteExcel.exists():
            deleteExcel.unlink()

        messages.success(request, "Aspirante eliminado correctamente")

    except Exception as e:
        messages.error(request, f"Error al eliminar aspirante: {e}")

    return redirect('consultas_instructor')
