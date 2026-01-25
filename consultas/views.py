from django.shortcuts import render, get_object_or_404, redirect
from django.http import Http404, FileResponse,  HttpResponse
from django.template.loader import render_to_string
from Cursos.models import (
    Usuario, Solicitud, Programaformacion, Horario, Modalidad,
    Departamentos, Municipios, Empresa, Programaespecial,
    Aspirantes, Caracterizacion, Tipoidentificacion, Estados, Ficha,
    Solicitudcoordinador, EstadosCoordinador, Area, Tipoempresa, Tiposolicitud, Usuariosasignados
)
# COnvertir ficha de caracterizacion a pdf
from weasyprint import HTML, CSS
# Sirve para Generar tokens, contraseñas y urls
import secrets
# Convertir todo a cadena
import string
import os
# Poder buscar la ruta para descargar el PDF
from django.conf import settings
# Importar las fechas
import datetime
# Libreria para usar el calendario
import calendar
# Crear copias
import shutil
# Importar mensajes
from django.contrib import messages
# Poder visualizar el excel
from openpyxl import load_workbook
# Importar decorador personalizado para el logueo
from Cursos.views import login_required_custom

from django.http import HttpResponseNotFound
# from xhtml2pdf import pisa
from io import BytesIO
from django.template.loader import get_template
# from io import BytesIO

# import pandas as pd

# Ver pdf aspirante en especifico

def showPdfApplicants(request, id, numDoc):
    # Verificar que el aspirante existe
    if not Aspirantes.objects.filter(numeroidentificacion=numDoc).exists():
        raise Http404("El aspirante no existe")

    pdf_path = os.path.join(
        settings.MEDIA_ROOT,
        'pdf',
        f'solicitud_{id}',
        f'{numDoc}.pdf'
    )

    if not os.path.exists(pdf_path):
        raise Http404("El archivo no existe")

    return FileResponse(
        open(pdf_path, 'rb'),
        content_type='application/pdf'
    )

# Ver pdf combinado
def viewCombinedPdf (request,  pdfFolder):

    pdfPath = os.path.join(
        settings.MEDIA_ROOT,
        'pdf',
        f'solicitud_{pdfFolder}',
        'combinado.pdf'
    )

    if not os.path.exists(pdfPath):
        raise Http404('El archivo no existe')

    return FileResponse(
        open(pdfPath, 'rb'),
        content_type='application/pdf'
    )

def showExcelApprentices(request, excelFolder ):
    # Nombre del archivo
    folder_name = f'formato_inscripcion_{excelFolder}.xlsx'
    ruta_archivo = os.path.join(settings.MEDIA_ROOT, "excel", folder_name)

    datos = []
    error = None

    try:
        # Cargar el archivo Excel
        wb = load_workbook(ruta_archivo)
        ws = wb.active  # primera hoja

        # Recorrer todas las filas y columnas
        for row in ws.iter_rows(values_only=True):
            fila = []
            for celda in row:
                fila.append("" if celda is None else str(celda))
            datos.append(fila)

    except Exception as e:
            error = f"Error al abrir el archivo: {e}"

    return render(request, "fichacaracterizacion/formato_inscripcion.html", {
        "datos": datos,
        "error": error
    })

def reviewedByInstructor(request, idSolicitud):
    user_id = request.session.get('user_id')

    usuario = Usuario.objects.select_related('rol').get(idusuario=user_id)
    id_rol = usuario.rol.idrol

    if id_rol != 1:
        return redirect('consultas_instructor')

    solicitud = get_object_or_404(Solicitud, idsolicitud=idSolicitud)
    solicitud.revisado = 1
    solicitud.save()

    return redirect('consultas_instructor')

def editApplicantData(request, idSolicitud, numDoc):

    user_id = request.session.get('user_id')
    if not user_id:
        raise Http404()

    usuario = Usuario.objects.select_related('rol').get(idusuario=user_id)

    # SOLO instructor puede editar
    if usuario.rol.idrol != 1:
        raise Http404()

    aspirante = get_object_or_404(
        Aspirantes,
        numeroidentificacion=numDoc,
        solicitudinscripcion=idSolicitud
    )

    tipoIdentificacion = Tipoidentificacion.objects.all()

    return render(request, 'forms/editApplicant.html', {
        "aspirante": aspirante,
        "tipoIdentificacion":tipoIdentificacion
    })

# =====================================================================
# Consultas dependiendo del rol
# =====================================================================
@login_required_custom
def consultas_todos(request):

    caracteres = string.ascii_letters + string.digits
    codigo = ''.join(secrets.choice(caracteres) for _ in range(5))

    user_id = request.session.get('user_id')
    usuario = Usuario.objects.select_related('rol').get(idusuario=user_id)
    id_rol = usuario.rol.idrol

    if id_rol == 1:
        layout = 'layout/layoutinstructor.html'
        rol_name = 'Instructor'
    elif id_rol == 2:
        layout = 'layout/layout_coordinador.html'
        rol_name = 'Coordinador'
    elif id_rol == 3:
        layout = 'layout/layout_funcionario.html'
        rol_name = 'Funcionario'
    elif id_rol == 4:
        layout = 'layout/layout_admin.html'
        rol_name = "Administrador"
    else:
        layout = 'layout/layout_admin.html'
        rol_name = "Desconocido"

    # Solicitudes según rol
    if id_rol == 2:  # Coordinador: solo solicitudes de instructores asignados
        hoy = datetime.date.today()
        primer_dia = datetime.date(hoy.year, hoy.month, 1)
        ultimo_dia = datetime.date(hoy.year, hoy.month, calendar.monthrange(hoy.year, hoy.month)[1])

        instructores_asignados = Usuariosasignados.objects.filter(
            idusuariocoordinador=usuario
        ).values_list('idinstructor', flat=True)

        solicitudes = Solicitud.objects.filter(
            fechasolicitud__range=(primer_dia, ultimo_dia),
            idusuario__in=instructores_asignados
        ).select_related('idusuario', 'idempresa').order_by('-fechasolicitud')

        solicitudes = solicitudes.filter(revisado=1)

    elif id_rol == 3:  # Funcionario: solicitudes aprobadas por coordinador
        solicitudes_aprobadas = Solicitudcoordinador.objects.filter(
            idestado__estado="Aprobado"
        ).values_list('idsolicitud', flat=True)

        solicitudes = Solicitud.objects.filter(
            idsolicitud__in=solicitudes_aprobadas
        ).select_related('idusuario', 'idempresa').order_by('-fechasolicitud')
    else:  # Instructor o Admin
        solicitudes = Solicitud.objects.select_related('idusuario', 'idempresa') \
            .filter(idusuario=user_id) \
            .order_by('-fechasolicitud')

    # Estados según rol
    if id_rol == 3:
        estado = Estados.objects.values('idestado', 'estados')
    elif id_rol == 2:
        estado = EstadosCoordinador.objects.values('id', 'estado')
    else:
        estado = Estados.objects.none()

    for solicitud in solicitudes:
        ultima_revision = Solicitudcoordinador.objects.filter(
            idsolicitud=solicitud
        ).select_related('usuario_revisador', 'usuario_solicitud', 'idestado').order_by('-fecha').first()

        if ultima_revision:
            solicitud.estado_coordinador = ultima_revision.idestado.estado
            solicitud.observacion_coordinador = ultima_revision.observacion or ''
        else:
            solicitud.estado_coordinador = None
            solicitud.observacion_coordinador = ''

        ficha = Ficha.objects.filter(idsolicitud=solicitud.idsolicitud).select_related('idestado', 'idusuario').first()
        if ficha:
            solicitud.estado_usuario = ficha.idestado.estados if ficha.idestado else None
            solicitud.observacion_usuario = ficha.observacion or ''
            solicitud.codigo_ficha = ficha.codigoficha or ''
        else:
            solicitud.estado_usuario = None
            solicitud.observacion_usuario = ''  # NO heredar observación del coordinador
            solicitud.codigo_ficha = ''


        solicitud.codigo_solicitud = solicitud.codigosolicitud

        if getattr(solicitud, 'idusuario', None):
            creador = solicitud.idusuario
            solicitud.visible_nombre = f"{creador.nombre} {creador.apellido}"
        else:
            solicitud.visible_nombre = "Creador desconocido"

        if id_rol == 2 and ultima_revision and ultima_revision.usuario_solicitud:
            solicitud.visible_nombre = f"{ultima_revision.usuario_solicitud.nombre} {ultima_revision.usuario_solicitud.apellido}"
        elif id_rol == 3 and ultima_revision and ultima_revision.idestado.estado == "Aprobado" and ultima_revision.usuario_revisador:
            solicitud.visible_nombre = f"{ultima_revision.usuario_revisador.nombre} {ultima_revision.usuario_revisador.apellido}"
        elif id_rol == 1:
            if ficha and ficha.idusuario:
                funcionario = ficha.idusuario
                solicitud.visible_nombre = f"{funcionario.nombre} {funcionario.apellido}"
            else:
                solicitud.visible_nombre = "Sin aprobación aún"

        solicitud.boton_ver_carta_disabled = (id_rol == 2 and solicitud.idempresa is None)
        solicitud.mostrar_boton_carta_funcionario = (id_rol == 3 and solicitud.idempresa is None)

        try:
            ruta_excel_funcionario = os.path.join(settings.MEDIA_ROOT, 'excel', f'formato_inscripcion_{solicitud.idsolicitud}.xlsx')
            solicitud.excel_funcionario_disponible = os.path.exists(ruta_excel_funcionario)

            carpeta_excel_sofia = os.path.join(
                settings.MEDIA_ROOT,
                'Funcionario',
                f"solicitud_{solicitud.idsolicitud}",
                'Masivos_sofia_plus'
            )
            ruta_excel_sofia = os.path.join(carpeta_excel_sofia, f'formato_inscripcion_{solicitud.idsolicitud}.xlsx')
            solicitud.excel_masivo_disponible = os.path.exists(ruta_excel_sofia)
        except Exception:
            solicitud.excel_funcionario_disponible = False
            solicitud.excel_masivo_disponible = False

    if id_rol in [1, 4]:
        for solicitud in solicitudes:
            aspirantes = Aspirantes.objects.select_related(
                'tipoidentificacion', 'idcaracterizacion'
            ).filter(solicitudinscripcion=solicitud.idsolicitud)
            solicitud.aspirantes = aspirantes

        # Por defecto NO mostrar nada
    solicitud.mostrar_pdf = False
    solicitud.mostrar_excel = False

    # SOLO ROL 1
    if id_rol == 1:
        # PDF combinado (creado con ID)
        ruta_pdf = os.path.join(
            settings.MEDIA_ROOT,
            'pdf',
            f'solicitud_{solicitud.idsolicitud}',
            'combinado.pdf'
        )

        # Excel (creado con ID)
        ruta_excel = os.path.join(
            settings.MEDIA_ROOT,
            'excel',
            f'formato_inscripcion_{solicitud.idsolicitud}.xlsx'
        )

        solicitud.mostrar_pdf = os.path.exists(ruta_pdf)
        solicitud.mostrar_excel = os.path.exists(ruta_excel)

    return render(request, "consultas/consultas_instructor.html", {
        "layout": layout,
        "rol": id_rol,
        "user": rol_name,
        'codigo': codigo,
        'solicitudes': solicitudes,
        'estado': estado
    })

# =====================================================================
# Reportes (página simple con layout según rol)
# =====================================================================
@login_required_custom
def reportes(request):
    """Página de Reportes: acceso SOLO para Coordinador (2) y Funcionario (3).

    Filtros disponibles:
    - instructor (Usuario con rol Instructor)
    - tipo de solicitud (Tiposolicitud: p.e. 'CAMPESENA' o 'REGULAR')
    - con/sin empresa
    - tipo de empresa (Tipoempresa)
    - estado de solicitud (EstadosCoordinador)
    - estado de ficha (Estados)
    - área (Area del Programaformacion)
    """
    user_id = request.session.get('user_id')
    usuario = get_object_or_404(Usuario.objects.select_related('rol'), idusuario=user_id)
    id_rol = usuario.rol.idrol

    # Restringir acceso: solo roles 2 y 3
    if id_rol not in (2, 3):
        messages.error(request, 'No tienes permisos para acceder a Reportes.')
        return redirect('consultas_instructor')

    # Layout y etiqueta de usuario según rol permitido
    layout = 'layout/layout_coordinador.html' if id_rol == 2 else 'layout/layout_funcionario.html'
    rol_name = 'Coordinador' if id_rol == 2 else 'Funcionario'

    # ---------------------------------
    # Parámetros de filtro (GET)
    # ---------------------------------
    instructor_id = request.GET.get('instructor') or ''
    tipo_solicitud_id = request.GET.get('tipo_solicitud') or ''
    con_empresa = request.GET.get('con_empresa') or ''  # '', 'si', 'no'
    tipo_empresa_id = request.GET.get('tipo_empresa') or ''
    estado_solicitud_id = request.GET.get('estado_solicitud') or ''  # EstadosCoordinador.id
    estado_ficha_id = request.GET.get('estado_ficha') or ''          # Estados.idestado
    area_id = request.GET.get('area') or ''

    # ---------------------------------
    # Query base
    # ---------------------------------
    solicitudes_qs = (
        Solicitud.objects
        .select_related(
            'idusuario', 'idtiposolicitud',
            'idempresa__idtipoempresa',
            'codigoprograma__idarea'
        )
        .order_by('-fechasolicitud')
    )

    # ---------------------------------
    # Aplicar filtros
    # ---------------------------------
    if instructor_id:
        solicitudes_qs = solicitudes_qs.filter(idusuario__idusuario=instructor_id)

    if tipo_solicitud_id:
        solicitudes_qs = solicitudes_qs.filter(idtiposolicitud__idtiposolicitud=tipo_solicitud_id)

    if con_empresa == 'si':
        solicitudes_qs = solicitudes_qs.filter(idempresa__isnull=False)
    elif con_empresa == 'no':
        solicitudes_qs = solicitudes_qs.filter(idempresa__isnull=True)

    if tipo_empresa_id:
        solicitudes_qs = solicitudes_qs.filter(idempresa__idtipoempresa__idtipoempresa=tipo_empresa_id)

    if estado_ficha_id:
        # Estado de Ficha (relación inversa a Ficha)
        solicitudes_qs = solicitudes_qs.filter(ficha__idestado__idestado=estado_ficha_id)

    if estado_solicitud_id:
        # Estado de Solicitud (revisión del coordinador)
        solicitudes_qs = solicitudes_qs.filter(solicitudcoordinador__idestado__id=estado_solicitud_id)

    if area_id:
        solicitudes_qs = solicitudes_qs.filter(codigoprograma__idarea__idarea=area_id)

    solicitudes_qs = solicitudes_qs.distinct()

    # ---------------------------------
    # Enriquecer con estados de ficha y solicitud (evitar N+1)
    # ---------------------------------
    solicitudes = list(solicitudes_qs)
    if solicitudes:
        ids = [s.idsolicitud for s in solicitudes]

        # Fichas por solicitud
        fichas = (
            Ficha.objects
            .filter(idsolicitud_id__in=ids)
            .select_related('idestado')
        )
        fichas_map = {f.idsolicitud_id: f for f in fichas}

        # Última revisión del coordinador por solicitud (si hubiese varias, tomamos la más reciente por fecha)
        revisiones = (
            Solicitudcoordinador.objects
            .filter(idsolicitud_id__in=ids)
            .select_related('idestado')
            .order_by('idsolicitud_id', '-fecha')
        )
        revisiones_map = {}
        for r in revisiones:
            if r.idsolicitud_id not in revisiones_map:
                revisiones_map[r.idsolicitud_id] = r

        # Asignar campos calculados a cada solicitud
        for s in solicitudes:
            f = fichas_map.get(s.idsolicitud)
            s.estado_ficha_nombre = f.idestado.estados if f and f.idestado else None
            s.codigo_ficha = f.codigoficha if f else None

            rv = revisiones_map.get(s.idsolicitud)
            s.estado_solicitud_nombre = rv.idestado.estado if rv and rv.idestado else None

            s.tipo_empresa_nombre = s.idempresa.idtipoempresa.tipoempresa if s.idempresa and s.idempresa.idtipoempresa else None
            s.area_nombre = s.codigoprograma.idarea.area if s.codigoprograma and s.codigoprograma.idarea else None

    # ---------------------------------
    # Métricas basadas en los filtros aplicados
    # ---------------------------------
    from collections import Counter

    tipos_cont = Counter()
    empresa_con = 0
    empresa_sin = 0
    estado_sol_cont = Counter()
    estado_ficha_cont = Counter()
    area_cont = Counter()
    tipo_empresa_cont = Counter()

    def normalizar_tipo_solicitud(valor):
        if not valor:
            return '—'
        v = str(valor).lower()
        return 'CampeSENA' if v in ('campesena', 'campesina') else valor

    for s in solicitudes:
        # Tipo de solicitud
        tipos_cont[normalizar_tipo_solicitud(getattr(getattr(s, 'idtiposolicitud', None), 'tiposolicitud', None))] += 1

        # Empresa
        if getattr(s, 'idempresa', None):
            empresa_con += 1
            tipo_empresa_cont[getattr(getattr(s, 'idempresa', None), 'idtipoempresa', None) and s.idempresa.idtipoempresa.tipoempresa or '—'] += 1
        else:
            empresa_sin += 1

        # Estados
        estado_sol_cont[s.estado_solicitud_nombre or '—'] += 1
        estado_ficha_cont[s.estado_ficha_nombre or 'Sin ficha'] += 1

        # Área
        area_cont[s.area_nombre or '—'] += 1

    def ordenar_items(counter_obj):
        return sorted(counter_obj.items(), key=lambda x: (-x[1], str(x[0])))

    metricas = {
        'total': len(solicitudes),
        'empresa': {
            'con': empresa_con,
            'sin': empresa_sin,
        },
    }

    metricas_tipos = ordenar_items(tipos_cont)
    metricas_estado_sol = ordenar_items(estado_sol_cont)
    metricas_estado_ficha = ordenar_items(estado_ficha_cont)
    metricas_area = ordenar_items(area_cont)
    metricas_tipo_empresa = ordenar_items(tipo_empresa_cont)

    # ---------------------------------
    # Catálogos para filtros
    # ---------------------------------
    instructores = Usuario.objects.filter(rol__idrol=1).order_by('nombre', 'apellido')
    tipos_solicitud = Tiposolicitud.objects.all().order_by('tiposolicitud')
    tipos_empresa = Tipoempresa.objects.all().order_by('tipoempresa')
    estados_ficha = Estados.objects.all().order_by('estados')
    estados_sol = EstadosCoordinador.objects.all().order_by('estado')
    areas = Area.objects.all().order_by('area')

    return render(request, 'reportes/reportes.html', {
        'layout': layout,
        'rol': id_rol,
        'user': rol_name,
        'title': 'Reportes',
        # resultados
        'solicitudes': solicitudes,
        # opciones
        'instructores': instructores,
        'tipos_solicitud': tipos_solicitud,
        'tipos_empresa': tipos_empresa,
        'estados_ficha': estados_ficha,
        'estados_solicitud': estados_sol,
        'areas': areas,
        # valores seleccionados
        'f_instructor': instructor_id,
        'f_tipo_solicitud': tipo_solicitud_id,
        'f_con_empresa': con_empresa,
        'f_tipo_empresa': tipo_empresa_id,
        'f_estado_solicitud': estado_solicitud_id,
        'f_estado_ficha': estado_ficha_id,
        'f_area': area_id,
        'total_resultados': len(solicitudes),
        # métricas
        'metricas': metricas,
        'metricas_tipos': metricas_tipos,
        'metricas_estado_solicitud': metricas_estado_sol,
        'metricas_estado_ficha': metricas_estado_ficha,
        'metricas_area': metricas_area,
        'metricas_tipo_empresa': metricas_tipo_empresa,
    })


# ===============================================================================
# Mostrar la ficha de caracterización
# ===============================================================================
@login_required_custom
def ficha_caracterizacion(request, solicitud_id):
    """
    Vista para mostrar la ficha de caracterización
    """
    # Obtener el usuario actual desde la sesión
    user_id = request.session.get('user_id')
    usuario_actual = get_object_or_404(Usuario.objects.select_related('rol'), idusuario=user_id)

    # Obtener la solicitud con todas las relaciones necesarias
    # 🔹 select_related solo con ForeignKey / OneToOne
    solicitud = get_object_or_404(
        Solicitud.objects.select_related(
            'codigoprograma',
            'idhorario',
            'idmodalidad',
            'codigomunicipio__codigodepartamento',
            'idusuario',
            'idempresa',
            'idespecial',
        ),
        idsolicitud=solicitud_id
    )

    # Obtener todas las variables que necesita el template
    programa = solicitud.codigoprograma
    horario = solicitud.idhorario
    modalidad = solicitud.idmodalidad
    municipio = solicitud.codigomunicipio
    departamento = municipio.codigodepartamento
    usuario = solicitud.idusuario
    empresa = solicitud.idempresa
    programa_especial = solicitud.idespecial
    ambiente = solicitud.ambiente  #Campo de texto, se usa directo

    # Definir layout según rol del usuario actual
    id_rol = usuario_actual.rol.idrol
    if id_rol == 1:
        layout = 'layout/layoutinstructor.html'
    elif id_rol == 2:
        layout = 'layout/layout_coordinador.html'
    elif id_rol == 3:
        layout = 'layout/layout_funcionario.html'
    elif id_rol == 4:
        layout = 'layout/layout_admin.html'
    else:
        layout = 'layout/layout_admin.html'

    # Contexto que se envía al template
    context = {
        'layout': layout,
        'rol': id_rol,
        'solicitud': solicitud,
        'programa': programa,
        'horario': horario,
        'modalidad': modalidad,
        'departamento': departamento,
        'municipio': municipio,
        'usuario': usuario,
        'empresa': empresa,
        'programa_especial': programa_especial,
        'ambiente': ambiente,  # Se pasa al template como texto
    }

    # Renderizar template con datos
    return render(request, 'fichacaracterizacion/fichacaracterizacion.html', context)

# ======================================================================
# Generar la ficha de caracterización en pdf
# ======================================================================
@login_required_custom
def ficha_caracterizacion_pdf(request, solicitud_id):
    """Genera un PDF de la ficha de caracterización usando WeasyPrint.
    Solo lo guarda en disco si el rol es 3 (funcionario). En otros roles, solo lo descarga sin guardar.
    """

    # Reutilizar la lógica de la vista HTML
    user_id = request.session.get('user_id')
    usuario_actual = get_object_or_404(
        Usuario.objects.select_related('rol'),
        idusuario=user_id
    )

    # 🔹 Quitar 'ambiente' porque es CharField (no relacional)
    solicitud = get_object_or_404(
        Solicitud.objects.select_related(
            'codigoprograma',
            'idhorario',
            'idmodalidad',
            'codigomunicipio__codigodepartamento',
            'idusuario',
            'idempresa',
            'idespecial',
        ),
        idsolicitud=solicitud_id
    )

    # Obtener variables necesarias para el template
    programa = solicitud.codigoprograma
    horario = solicitud.idhorario
    modalidad = solicitud.idmodalidad
    municipio = solicitud.codigomunicipio
    departamento = municipio.codigodepartamento
    usuario = solicitud.idusuario
    empresa = solicitud.idempresa
    programa_especial = solicitud.idespecial
    ambiente = solicitud.ambiente  # ✅ Se accede directo porque es CharField

    # Definir layout según rol
    id_rol = usuario_actual.rol.idrol
    if id_rol == 1:
        layout = 'layout/layoutinstructor.html'
    elif id_rol == 2:
        layout = 'layout/layout_coordinador.html'
    elif id_rol == 3:
        layout = 'layout/layoutfuncionario.html'
    elif id_rol == 4:
        layout = 'layout/layout_admin.html'
    else:
        layout = 'layout/layout_admin.html'

    # Contexto enviado al template
    context = {
        'layout': layout,
        'rol': id_rol,
        'solicitud': solicitud,
        'programa': programa,
        'horario': horario,
        'modalidad': modalidad,
        'departamento': departamento,
        'municipio': municipio,
        'usuario': usuario,
        'empresa': empresa,
        'programa_especial': programa_especial,
        'ambiente': ambiente,  # Se pasa como texto
        'pdf_mode': True,
    }

    # Renderizar HTML con contexto
    html_string = render_to_string('fichacaracterizacion/fichacaracterizacion.html', context)

    # Ruta al CSS
    css_path = os.path.join(settings.BASE_DIR, 'Cursos', 'static', 'css', 'ficha-caracterizacion.css')

    html = HTML(string=html_string, base_url=request.build_absolute_uri('/'))
    css = CSS(filename=css_path)

    pdf_bytes = html.write_pdf(stylesheets=[css])

    # Solo si el rol es 3, guardar en disco
    if int(id_rol) == 3:
        folder_name = f"solicitud_{solicitud.idsolicitud}"
        carpeta_destino = os.path.join(settings.MEDIA_ROOT, 'funcionario', folder_name)
        os.makedirs(carpeta_destino, exist_ok=True)

        filename_pdf = f"ficha_caracterizacion_{solicitud.idsolicitud}.pdf"
        ruta_guardado = os.path.join(carpeta_destino, filename_pdf)

        # Guardar archivo en disco
        with open(ruta_guardado, 'wb') as f:
            f.write(pdf_bytes)

        # Abrir el archivo guardado para descargarlo
        response_file = open(ruta_guardado, 'rb')
    else:
        # Para otros roles: no guardar en disco, usar los bytes directamente
        response_file = BytesIO(pdf_bytes)  # 🔹 Usar BytesIO para respuesta directa

    # Descargar el archivo
    return FileResponse(
        response_file,
        as_attachment=True,
        filename='ficha_caracterizacion.pdf',
        content_type='application/pdf'
    )

# ========================================================================
# Descargar el PDF combinado de los aspirantes
# ========================================================================
@login_required_custom
def descargar_pdf(request, id, idrol):
    folder_name = f"solicitud_{id}"

    # Ruta del archivo original
    buscar_pdf = os.path.join(settings.MEDIA_ROOT, 'pdf', folder_name, 'combinado.pdf')

    if not os.path.exists(buscar_pdf):
        raise Http404("PDF no encontrado")

    # Si el rol es 3 = funcionario crear una copia del archivo
    if int(idrol) == 3:
        # Ruta de destino donde se guardará una copia
        carpeta_destino = os.path.join(settings.MEDIA_ROOT, 'Funcionario',  folder_name)
        os.makedirs(carpeta_destino, exist_ok=True)

        guardar_pdf = os.path.join(carpeta_destino, 'combinado.pdf')

        # Copiar el archivo original al nuevo destino
        shutil.copy2(buscar_pdf, guardar_pdf)

    # Descargar el archivo original
    return FileResponse(
        open(buscar_pdf, 'rb'),
        as_attachment=True,
        filename='Documentos_aspirantes.pdf',
        content_type='application/pdf'
    )

# =======================================================================
# Descargar excel como funcionario
# =======================================================================
@login_required_custom
def descargar_excel(request, id, idrol):
    folder_name = f"solicitud_{id}"

    buscar_excel = os.path.join(settings.MEDIA_ROOT, 'excel', f'formato_inscripcion_{id}.xlsx')

    if not os.path.exists(buscar_excel):
        raise Http404("Excel no encontrado")

    # Si el rol es 3 = funcionario crear una copia del archivo
    if int(idrol) == 3:
        # Ruta de destino donde se guardará una copia
        carpeta_destino = os.path.join(settings.MEDIA_ROOT, 'Funcionario', folder_name)
        os.makedirs(carpeta_destino, exist_ok=True)

        guardar_excel = os.path.join(carpeta_destino, f'formato_inscripcion_{id}.xlsx')

        # Copiar el archivo original al nuevo destino
        shutil.copy2(buscar_excel, guardar_excel)

    # Descargar el archivo original
    return FileResponse(
        open(buscar_excel, 'rb'),
        as_attachment=True,
        filename='Formato_inscripcion.xlsx',
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# ========================================================================
# Descargar la carta de solicitud de la empresa
# ========================================================================
@login_required_custom
def descargar_carta(request, id, idrol):
    # ============================
    # Buscar la solicitud
    # ============================
    solicitud = get_object_or_404(Solicitud, idsolicitud=id)

    # ============================
    # Determinar qué PDF usar
    # ============================
    if solicitud.idempresa:  # Empresa existe
        # Usar el ID de la solicitud en lugar del NIT para la estructura de carpetas
        folder_name = f"carta_{solicitud.idsolicitud}"
        ruta_pdf = os.path.join(settings.MEDIA_ROOT, 'Cartas_de_solicitud', folder_name, f'carta_{solicitud.idsolicitud}.pdf')
    else:  # Empresa nula → usar carta interna generada
        folder_name = f"carta_{solicitud.idsolicitud}"
        ruta_pdf = os.path.join(settings.MEDIA_ROOT, 'Cartas_de_solicitud', folder_name, f'{folder_name}.pdf')

    # ============================
    # Verificar si el archivo existe
    # ============================
    if not os.path.exists(ruta_pdf):
        raise Http404("PDF no encontrado")

    # ============================
    # Si el rol es 3 = Funcionario, crear copia del archivo
    # ============================
    if int(idrol) == 3:
        carpeta_destino = os.path.join(settings.MEDIA_ROOT, 'Funcionario', f'solicitud_{solicitud.idsolicitud}')
        os.makedirs(carpeta_destino, exist_ok=True)

        # Ruta completa del archivo copia
        generar_copia = os.path.join(carpeta_destino, os.path.basename(ruta_pdf))

        # Copiar el archivo original al nuevo destino
        shutil.copy2(ruta_pdf, generar_copia)

    # ============================
    # Descargar el archivo
    # ============================
    return FileResponse(
        open(ruta_pdf, 'rb'),
        as_attachment=True,
        filename='Carta_solicitud.pdf',
        content_type='application/pdf'
    )

# ===========================================
# Funcionario respuestas a las solicitudes
# ===========================================
@login_required_custom
def revision_fichas(request, id):

    # ----------------------------
    # Buscar la solicitud
    # ----------------------------
    solicitud = get_object_or_404(Solicitud, idsolicitud=id)

    # ----------------------------
    # Obtener el id del usuario en sesión
    # ----------------------------
    user_id = request.session.get('user_id')

    # ----------------------------
    # Obtener el usuario y su rol
    # ----------------------------
    usuario = Usuario.objects.select_related('rol').get(idusuario=user_id)
    id_rol = usuario.rol.idrol

    # ----------------------------
    # Definir layout según rol del usuario actual
    # ----------------------------
    if id_rol == 1:
        layout = 'layout/layoutinstructor.html'
    elif id_rol == 2:
        layout = 'layout/layout_coordinador.html'
    elif id_rol == 3:
        layout = 'layout/layout_funcionario.html'
    elif id_rol == 4:
        layout = 'layout/layout_admin.html'
    else:
        layout = 'layout/layout_admin.html'

    # ----------------------------
    # Validar envío de formulario
    # ----------------------------
    if request.method == "POST":
        try:
            # ----------------------------
            # Capturar datos enviados
            # ----------------------------
            estados = request.POST.get('estado')
            numero_solicitud = request.POST.get('codigo_solicitud')
            numero_ficha = request.POST.get('codigo_ficha')

            numero_solicitud = int(numero_solicitud) if numero_solicitud else None
            numero_ficha = int(numero_ficha) if numero_ficha else None

            observacion = request.POST.get('observacion')
            nuevo_archivo = request.FILES.get('actualizar_excel', None)

            # ----------------------------
            # Obtener objetos relacionados
            # ----------------------------
            id_estado = Estados.objects.get(idestado=estados)
            creado_por = solicitud.idusuario

            # ----------------------------
            # Buscar si ya existe ficha para esta solicitud
            # ----------------------------
            ficha = Ficha.objects.filter(idsolicitud=solicitud).first()

            if ficha:
                # ----------------------------
                # Actualizar ficha existente
                # ----------------------------
                if numero_ficha is not None and numero_ficha != "":
                    ficha.codigoficha = numero_ficha  # solo si se envía valor nuevo

                ficha.idestado = id_estado
                ficha.observacion = observacion
                ficha.save()
            else:
                # ----------------------------
                # Crear nueva ficha si no existía
                # ----------------------------
                Ficha.objects.create(
                    codigoficha=numero_ficha if numero_ficha else None,
                    idsolicitud=solicitud,
                    idestado=id_estado,
                    idusuario=creado_por,
                    observacion=observacion,
                )

            # ----------------------------
            # Actualizar la solicitud con el número/código
            # ----------------------------
            if numero_solicitud is not None and numero_solicitud != "":
                solicitud.codigosolicitud = numero_solicitud  # solo si se envía valor nuevo
                solicitud.save()
            else:
                # Si no se envía, mantiene el valor existente
                pass

            # ----------------------------
            # Guardar Excel si se envía uno nuevo
            # ----------------------------
            if nuevo_archivo:
                carpeta_almacenar = f"solicitud_{id}"
                excel_archivo = f"formato_inscripcion_{id}.xlsx"
                carpeta_excel = os.path.join(
                    settings.MEDIA_ROOT,
                    'Funcionario',
                    carpeta_almacenar,
                    'Masivos_sofia_plus'
                )
                os.makedirs(carpeta_excel, exist_ok=True)

                ruta_excel = os.path.join(carpeta_excel, excel_archivo)
                with open(ruta_excel, 'wb+') as destino:
                    for chunk in nuevo_archivo.chunks():
                        destino.write(chunk)
            else:
                # Si no se envía archivo nuevo, mantiene el anterior
                pass

            # ----------------------------
            # Mensaje de éxito
            # ----------------------------
            messages.success(request, 'Haz enviado respuesta a esta solicitud')
            return redirect('consultas_instructor')

        except Exception as e:
            # ----------------------------
            # Mensaje de error en caso de excepción
            # ----------------------------
            messages.error(request, f'Error al enviar respuesta: {e}')
            return redirect('consultas_instructor')

    # ----------------------------
    # Renderizar plantilla si no hay envío de formulario
    # ----------------------------
    return render(request, 'consultas/consultas_instructor.html', {
        'layout': layout,
        'messages': messages
    })


# ===============================================================
# Descargar el formato generado por sofia plus
# ===============================================================
@login_required_custom
def descargar_excel_ficha(request, id):
    # Nombre del archivo esperado
    excel_archivo = f"formato_inscripcion_{id}.xlsx"

    # Ruta completa donde debe estar almacenado el archivo
    carpeta_excel = os.path.join(settings.MEDIA_ROOT, 'Funcionario', f"solicitud_{id}", 'Masivos_sofia_plus')
    directorio_excel = os.path.join(carpeta_excel, excel_archivo)

    # Validar si el archivo existe antes de descargar
    if os.path.exists(directorio_excel):
        with open(directorio_excel, 'rb') as archivo:
            response = HttpResponse(
                archivo.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response['Content-Disposition'] = f'attachment; filename={excel_archivo}'
            return response
    else:
        # Si no existe, mostrar un error
        raise Http404("Aun no se ha subido el excel.")

@login_required_custom
def revision_coordinador(request, id_solicitud):
    # ============================
    # Buscar la solicitud
    # ============================
    solicitud = get_object_or_404(Solicitud, idsolicitud=id_solicitud)

    # ============================
    # Obtener el usuario en sesión (el que revisa)
    # ============================
    user_id = request.session.get('user_id')
    usuario_revisador = Usuario.objects.select_related('rol').get(idusuario=user_id)
    id_rol = usuario_revisador.rol.idrol

    # ============================
    # Definir layout según rol
    # ============================
    if id_rol == 1:
        layout = 'layout/layoutinstructor.html'
    elif id_rol == 2:
        layout = 'layout/layout_coordinador.html'
    elif id_rol == 3:
        layout = 'layout/layout_funcionario.html'
    elif id_rol == 4:
        layout = 'layout/layout_admin.html'
    else:
        layout = 'layout/layout_admin.html'

    # ============================
    # Procesar POST
    # ============================
    if request.method == "POST":
        try:
            # ----------------------------
            # Capturar datos del formulario
            # ----------------------------
            estado_id = request.POST.get('estado')
            observacion = request.POST.get('observacion')

            if not estado_id:
                messages.error(request, 'Debe seleccionar un estado')
                return redirect('consultas_instructor')

            # ----------------------------
            # Obtener objeto estado
            # ----------------------------
            estado_obj = EstadosCoordinador.objects.get(id=estado_id)

            # ----------------------------
            # Usuario que creó la solicitud
            # ----------------------------
            usuario_solicitud = solicitud.idusuario

            # ----------------------------
            # Buscar si ya existe una revisión previa para esta solicitud
            # ----------------------------
            revision, creada = Solicitudcoordinador.objects.update_or_create(
                idsolicitud=solicitud,
                defaults={
                    "usuario_solicitud": usuario_solicitud,          # usuario que creó la solicitud
                    "usuario_revisador": usuario_revisador,  # coordinador que revisa
                    "idestado": estado_obj,
                    "observacion": observacion,
                    "fecha": datetime.date.today(),
                }
            )

            # ============================
            # Generar carta PDF internamente SOLO SI la revisión fue exitosa
            # ============================
            try:
                if solicitud.idempresa is None:
                    # Renderizar template HTML
                    template = get_template('fichacaracterizacion/carta_coordinador.html')
                    contexto = {
                        'solicitud': solicitud,
                        'usuario': usuario_revisador,
                    }
                    html_string = template.render(contexto)

                    # Crear carpeta para guardar PDF
                    folder_archive_name = f'carta_{solicitud.idsolicitud}'
                    carpeta = os.path.join(settings.MEDIA_ROOT, 'Cartas_de_solicitud', folder_archive_name)
                    os.makedirs(carpeta, exist_ok=True)

                    # Ruta del PDF
                    archivo_pdf = os.path.join(carpeta, f'{folder_archive_name}.pdf')

                    # Crear PDF con WeasyPrint
                    html = HTML(string=html_string)
                    html.write_pdf(target=archivo_pdf)

                    print(f"PDF generado y guardado en {archivo_pdf}")

            except Exception as pdf_error:
                print(f"Error generando PDF para solicitud {solicitud.idsolicitud}: {pdf_error}")

            # ============================
            # Mensajes de éxito
            # ============================
            if creada:
                messages.success(request, 'Solicitud revisada correctamente (nueva revisión creada)')
            else:
                messages.success(request, 'Solicitud revisada correctamente (revisión actualizada)')

            return redirect('consultas_instructor')

        except Exception as e:
            messages.error(request, f'Error al enviar respuesta: {e}')
            return redirect('consultas_instructor')

    # ============================
    # Si no es POST, renderizar la plantilla
    # ============================
    return render(request, 'consultas/consultas_instructor.html', {
        'layout': layout,
        'messages': messages
    })



@login_required_custom
def ver_formato_inscripcion(request, id_solicitud):
    # Nombre del archivo
    folder_name = f'formato_inscripcion_{id_solicitud}.xlsx'
    ruta_archivo = os.path.join(settings.MEDIA_ROOT, "excel", folder_name)

    datos = []
    error = None

    try:
        # Cargar el archivo Excel
        wb = load_workbook(ruta_archivo)
        ws = wb.active  # primera hoja

        # Recorrer todas las filas y columnas
        for row in ws.iter_rows(values_only=True):
            fila = []
            for celda in row:
                fila.append("" if celda is None else str(celda))
            datos.append(fila)

    except Exception as e:
            error = f"Error al abrir el archivo: {e}"

    return render(request, "fichacaracterizacion/formato_inscripcion.html", {
        "datos": datos,
        "error": error
    })

@login_required_custom
def ver_pdf_aspirantes(request, id_solicitud):
    folder_name = f'solicitud_{id_solicitud}'
    ruta_pdf = f"{settings.MEDIA_URL}pdf/{folder_name}/combinado.pdf"
    return redirect(ruta_pdf)

@login_required_custom
def ver_pdf_carta(request, id_solicitud):
    # Buscar la solicitud
    solicitud = get_object_or_404(Solicitud, idsolicitud=id_solicitud)

    # Tomar el ID de la solicitud
    idsolicitud = solicitud.idsolicitud

    # Validar que exista la solicitud
    if not idsolicitud:
        return HttpResponseNotFound("Esta solicitud no existe o no tiene un ID válido.")

    # Nombre de la carpeta y archivo
    folder_name = f"carta_{idsolicitud}"
    relative_path = f"Cartas_de_solicitud/{folder_name}/{folder_name}.pdf"
    absolute_path = os.path.join(settings.MEDIA_ROOT, relative_path)

    # Verificar si el archivo existe
    if os.path.exists(absolute_path):
        ruta_pdf = f"{settings.MEDIA_URL}{relative_path}"
        return redirect(ruta_pdf)
    else:
        return HttpResponseNotFound("El archivo PDF no existe para esta solicitud.")
